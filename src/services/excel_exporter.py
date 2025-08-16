"""
Excel export service for RemotelyX job postings.
Handles data export to Excel with formatting and backup functionality.
"""

import os
import shutil
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from ..core.config import get_settings
from ..models.job_posting import JobPosting, JobStatistics


class ExcelExporter:
    """Handles Excel export operations for job postings."""
    
    def __init__(self):
        self.settings = get_settings()
        self.ensure_directories()
    
    def ensure_directories(self) -> None:
        """Ensure required directories exist."""
        Path(self.settings.excel_file_path).parent.mkdir(parents=True, exist_ok=True)
        Path(self.settings.excel_backup_dir).mkdir(parents=True, exist_ok=True)
    
    def export_jobs_to_excel(self, jobs: List[JobPosting]) -> Dict[str, Any]:
        """Export job postings to Excel with formatting."""
        try:
            logger.info(f"Exporting {len(jobs)} jobs to Excel")
            
            # Convert jobs to DataFrame
            df = self._jobs_to_dataframe(jobs)
            
            # Create Excel file with formatting
            self._create_excel_file(df)
            
            # Generate statistics
            stats = self._generate_export_stats(jobs)
            
            logger.success(f"Successfully exported {len(jobs)} jobs to Excel")
            return {
                "success": True,
                "exported": len(jobs),
                "file_path": self.settings.excel_file_path,
                "statistics": stats
            }
            
        except Exception as e:
            logger.error(f"Failed to export jobs to Excel: {e}")
            return {
                "success": False,
                "error": str(e),
                "exported": 0
            }
    
    def _jobs_to_dataframe(self, jobs: List[JobPosting]) -> pd.DataFrame:
        """Convert job postings to pandas DataFrame."""
        data = []
        
        for job in jobs:
            job_dict = {
                "Job ID": job.job_id or "",
                "Title": job.title,
                "Company": job.company,
                "Location": job.location,
                "Employment Type": job.employment_type.value if job.employment_type else "",
                "Seniority Level": job.seniority_level.value if job.seniority_level else "",
                "Work Mode": job.work_mode.value if job.work_mode else "",
                "Salary Min": job.salary_min or "",
                "Salary Max": job.salary_max or "",
                "Salary Currency": job.salary_currency,
                "Salary Period": job.salary_period,
                "Skills": ", ".join(job.skills) if job.skills else "",
                "Tags": ", ".join(job.tags) if job.tags else "",
                "Description": job.description[:500] + "..." if len(job.description) > 500 else job.description,
                "Job URL": job.job_url,
                "Posting Date": job.posting_date.strftime("%Y-%m-%d") if job.posting_date else "",
                "Scraped At": job.scraped_at.strftime("%Y-%m-%d %H:%M:%S"),
                "Processed": "Yes" if job.processed else "No",
                "Enriched": "Yes" if job.enriched else "No",
                "Created At": job.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "Updated At": job.updated_at.strftime("%Y-%m-%d %H:%M:%S")
            }
            data.append(job_dict)
        
        return pd.DataFrame(data)
    
    def _create_excel_file(self, df: pd.DataFrame) -> None:
        """Create Excel file with proper formatting."""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Postings"
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Apply formatting
        self._apply_excel_formatting(ws, df)
        
        # Save file
        wb.save(self.settings.excel_file_path)
    
    def _apply_excel_formatting(self, ws, df: pd.DataFrame) -> None:
        """Apply formatting to Excel worksheet."""
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_export_stats(self, jobs: List[JobPosting]) -> Dict[str, Any]:
        """Generate statistics for the exported data."""
        if not jobs:
            return {}
        
        stats = {
            "total_jobs": len(jobs),
            "companies": list(set(job.company for job in jobs if job.company)),
            "locations": list(set(job.location for job in jobs if job.location)),
            "seniority_levels": {},
            "work_modes": {},
            "employment_types": {},
            "skills_frequency": {},
            "salary_stats": {}
        }
        
        # Count seniority levels
        for job in jobs:
            if job.seniority_level:
                level = job.seniority_level.value
                stats["seniority_levels"][level] = stats["seniority_levels"].get(level, 0) + 1
        
        # Count work modes
        for job in jobs:
            if job.work_mode:
                mode = job.work_mode.value
                stats["work_modes"][mode] = stats["work_modes"].get(mode, 0) + 1
        
        # Count employment types
        for job in jobs:
            if job.employment_type:
                emp_type = job.employment_type.value
                stats["employment_types"][emp_type] = stats["employment_types"].get(emp_type, 0) + 1
        
        # Count skills
        for job in jobs:
            for skill in job.skills:
                stats["skills_frequency"][skill] = stats["skills_frequency"].get(skill, 0) + 1
        
        # Salary statistics
        salaries = [job.salary_min for job in jobs if job.salary_min is not None]
        if salaries:
            stats["salary_stats"] = {
                "min": min(salaries),
                "max": max(salaries),
                "avg": sum(salaries) / len(salaries),
                "count": len(salaries)
            }
        
        return stats
    
    def create_backup(self) -> Dict[str, Any]:
        """Create a backup of the current Excel file."""
        try:
            if not os.path.exists(self.settings.excel_file_path):
                return {
                    "success": False,
                    "error": "No Excel file to backup"
                }
            
            # Create backup filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"job_postings_backup_{timestamp}.xlsx"
            backup_path = os.path.join(self.settings.excel_backup_dir, backup_filename)
            
            # Copy file
            shutil.copy2(self.settings.excel_file_path, backup_path)
            
            logger.success(f"Created backup: {backup_path}")
            return {
                "success": True,
                "backup_path": backup_path,
                "timestamp": timestamp
            }
            
        except Exception as e:
            logger.error(f"Failed to create backup: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def get_excel_statistics(self) -> Dict[str, Any]:
        """Get statistics from the current Excel file."""
        try:
            if not os.path.exists(self.settings.excel_file_path):
                return {}
            
            # Read Excel file
            df = pd.read_excel(self.settings.excel_file_path)
            
            stats = {
                "total_rows": len(df),
                "last_modified": datetime.fromtimestamp(
                    os.path.getmtime(self.settings.excel_file_path)
                ).strftime("%Y-%m-%d %H:%M:%S"),
                "file_size_mb": round(os.path.getsize(self.settings.excel_file_path) / (1024 * 1024), 2)
            }
            
            # Column statistics
            if not df.empty:
                stats["columns"] = list(df.columns)
                stats["companies"] = df["Company"].dropna().unique().tolist()
                stats["locations"] = df["Location"].dropna().unique().tolist()
                
                # Skills analysis
                if "Skills" in df.columns:
                    all_skills = []
                    for skills_str in df["Skills"].dropna():
                        if skills_str:
                            skills = [s.strip() for s in skills_str.split(",")]
                            all_skills.extend(skills)
                    
                    from collections import Counter
                    skills_counter = Counter(all_skills)
                    stats["top_skills"] = skills_counter.most_common(20)
            
            return stats
            
        except Exception as e:
            logger.error(f"Failed to get Excel statistics: {e}")
            return {}
    
    def list_backups(self) -> List[Dict[str, Any]]:
        """List all available backups."""
        try:
            backup_files = []
            backup_dir = Path(self.settings.excel_backup_dir)
            
            if not backup_dir.exists():
                return []
            
            for file_path in backup_dir.glob("job_postings_backup_*.xlsx"):
                backup_files.append({
                    "filename": file_path.name,
                    "path": str(file_path),
                    "size_mb": round(file_path.stat().st_size / (1024 * 1024), 2),
                    "created_at": datetime.fromtimestamp(file_path.stat().st_ctime).strftime("%Y-%m-%d %H:%M:%S")
                })
            
            # Sort by creation time (newest first)
            backup_files.sort(key=lambda x: x["created_at"], reverse=True)
            return backup_files
            
        except Exception as e:
            logger.error(f"Failed to list backups: {e}")
            return []
    
    def restore_backup(self, backup_filename: str) -> Dict[str, Any]:
        """Restore from a backup file."""
        try:
            backup_path = os.path.join(self.settings.excel_backup_dir, backup_filename)
            
            if not os.path.exists(backup_path):
                return {
                    "success": False,
                    "error": "Backup file not found"
                }
            
            # Create backup of current file before restoring
            if os.path.exists(self.settings.excel_file_path):
                self.create_backup()
            
            # Restore backup
            shutil.copy2(backup_path, self.settings.excel_file_path)
            
            logger.success(f"Restored backup: {backup_filename}")
            return {
                "success": True,
                "restored_file": self.settings.excel_file_path
            }
            
        except Exception as e:
            logger.error(f"Failed to restore backup: {e}")
            return {
                "success": False,
                "error": str(e)
            }


# Global Excel exporter instance
excel_exporter = ExcelExporter() 